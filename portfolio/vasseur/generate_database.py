"""
Maison Vasseur -- Wholesale & Franchise Database Generator (2024-2025)

Generates a synthetic but internally-consistent EMEA wholesale/franchise
dataset for a fictitious luxury footwear maison, targeting ~CHF 200M/year
in total sell-in value across 12 accounts x 16 products x 24 months.

Output: a single .xlsx with four flat tables (Accounts, Products,
Transactions_Monthly, Transactions_Weekly), ready for Power Query import.
No formulas, no pivots -- raw source data only.
"""

import random
import datetime as dt
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

random.seed(42)

TARGET_ANNUAL_REVENUE = 200_000_000  # CHF, sanity-check target per year

# ---------------------------------------------------------------------------
# 1. ACCOUNTS
# ---------------------------------------------------------------------------
# share            = fraction of TARGET_ANNUAL_REVENUE this account represents (avg of 2024/2025)
# trend            = 2025 vs 2024 revenue multiplier (sell-in follows this)
# st_start/st_end  = monthly sell-through rate (of available stock) at month 1 vs month 24
# return_rate      = fraction of sell-out returned in an average month
# carried          = number of the 16 products this account carries

ACCOUNTS = [
    dict(id="ACC-01", name="Hartwell & Vane", region="UK and Ireland", country="United Kingdom",
         acct_type="Wholesale", share=0.13, trend=1.05, st_start=0.40, st_end=0.44, return_rate=0.030, carried=14),
    dict(id="ACC-02", name="Belgrave Maison", region="UK and Ireland", country="United Kingdom",
         acct_type="Franchise", share=0.05, trend=1.22, st_start=0.24, st_end=0.34, return_rate=0.030, carried=9),
    dict(id="ACC-03", name="Nordkapp Trading House", region="Nordics and Baltics", country="Sweden",
         acct_type="Wholesale", share=0.07, trend=1.04, st_start=0.35, st_end=0.37, return_rate=0.025, carried=12),
    dict(id="ACC-04", name="Stureplan Atelier", region="Nordics and Baltics", country="Sweden",
         acct_type="Franchise", share=0.04, trend=0.85, st_start=0.26, st_end=0.18, return_rate=0.040, carried=8),
    dict(id="ACC-05", name="Alpenrose Schuhhaus", region="DACH", country="Germany",
         acct_type="Wholesale", share=0.15, trend=1.06, st_start=0.42, st_end=0.46, return_rate=0.025, carried=15),
    dict(id="ACC-06", name="Konigsallee Boutique", region="DACH", country="Germany",
         acct_type="Franchise", share=0.05, trend=1.00, st_start=0.30, st_end=0.09, return_rate=0.110, carried=9),
    dict(id="ACC-07", name="Le Comptoir de la Mode", region="France and Benelux", country="France",
         acct_type="Wholesale", share=0.14, trend=1.08, st_start=0.38, st_end=0.42, return_rate=0.030, carried=14),
    dict(id="ACC-08", name="Maison Doree", region="France and Benelux", country="France",
         acct_type="Franchise", share=0.06, trend=1.10, st_start=0.27, st_end=0.33, return_rate=0.030, carried=10),
    dict(id="ACC-09", name="Casa Serrano", region="Iberia", country="Spain",
         acct_type="Wholesale", share=0.08, trend=1.05, st_start=0.33, st_end=0.35, return_rate=0.030, carried=12),
    dict(id="ACC-10", name="Boutique Almendra", region="Iberia", country="Spain",
         acct_type="Franchise", share=0.05, trend=0.95, st_start=0.28, st_end=0.09, return_rate=0.130, carried=8),
    dict(id="ACC-11", name="Casa Lombardi", region="Southern Europe", country="Italy",
         acct_type="Wholesale", share=0.14, trend=1.12, st_start=0.37, st_end=0.43, return_rate=0.025, carried=14),
    dict(id="ACC-12", name="Atelier del Sud", region="Southern Europe", country="Italy",
         acct_type="Franchise", share=0.04, trend=0.92, st_start=0.22, st_end=0.17, return_rate=0.045, carried=8),
]
assert abs(sum(a["share"] for a in ACCOUNTS) - 1.0) < 1e-9

# ---------------------------------------------------------------------------
# 2. PRODUCTS
# ---------------------------------------------------------------------------
# category: Women's / Men's / Leather Goods
# season tag = collection the SKU launched under (core SS24/FW24 items sell
# across the whole 24-month window; SS25/FW25 items are newer launches and
# are only active from their launch month onward)
# popularity: relative sell-through weight used to split revenue across a
# carrying account's assortment

PRODUCTS = [
    dict(id="PRD-01", name="Etoile Pump 100", category="Women's Footwear", season="SS24", wholesale=620, popularity=1.40),
    dict(id="PRD-02", name="Vasseur Oxford", category="Men's Footwear", season="FW24", wholesale=590, popularity=1.10),
    dict(id="PRD-03", name="Cambon Slingback", category="Women's Footwear", season="SS24", wholesale=640, popularity=1.20),
    dict(id="PRD-04", name="Rive Loafer", category="Men's Footwear", season="FW24", wholesale=480, popularity=1.15),
    dict(id="PRD-05", name="Voltige Sandal", category="Women's Footwear", season="SS25", wholesale=560, popularity=1.00),
    dict(id="PRD-06", name="Nocturne Derby", category="Men's Footwear", season="FW24", wholesale=560, popularity=1.05),
    dict(id="PRD-07", name="Lumiere Sneaker", category="Women's Footwear", season="SS25", wholesale=490, popularity=1.10),
    dict(id="PRD-08", name="Faubourg Sneaker", category="Men's Footwear", season="FW25", wholesale=470, popularity=1.00),
    dict(id="PRD-09", name="Aurore Ankle Boot", category="Women's Footwear", season="FW24", wholesale=710, popularity=0.90),
    dict(id="PRD-10", name="Bastille Boot", category="Men's Footwear", season="FW25", wholesale=690, popularity=0.80),
    dict(id="PRD-11", name="Soiree Mule", category="Women's Footwear", season="SS25", wholesale=600, popularity=0.90),
    dict(id="PRD-12", name="Cambon Card Holder", category="Leather Goods", season="SS24", wholesale=190, popularity=1.00),
    dict(id="PRD-13", name="Voltige Belt", category="Leather Goods", season="FW24", wholesale=340, popularity=0.90),
    dict(id="PRD-14", name="Etoile Zip Wallet", category="Leather Goods", season="SS24", wholesale=260, popularity=0.95),
    dict(id="PRD-15", name="Rive Crossbody Pouch", category="Leather Goods", season="SS25", wholesale=430, popularity=0.85),
    dict(id="PRD-16", name="Lumiere Passport Cover", category="Leather Goods", season="FW25", wholesale=210, popularity=0.90),
]
RETAIL_MULTIPLIER = {"Women's Footwear": 2.5, "Men's Footwear": 2.45, "Leather Goods": 2.35}
for p in PRODUCTS:
    p["retail"] = int(round(p["wholesale"] * RETAIL_MULTIPLIER[p["category"]] / 10.0)) * 10

FOOTWEAR_ORDER = [p["id"] for p in PRODUCTS if p["category"] != "Leather Goods"]  # PRD-01..PRD-11 in this order
LEATHER_ORDER = [p["id"] for p in PRODUCTS if p["category"] == "Leather Goods"]   # PRD-12..PRD-16 in this order
PRODUCTS_BY_ID = {p["id"]: p for p in PRODUCTS}

# ---------------------------------------------------------------------------
# 3. TIME DIMENSIONS
# ---------------------------------------------------------------------------
MONTHS = [(y, m) for y in (2024, 2025) for m in range(1, 13)]  # 24 months
MONTH_INDEX = {ym: i for i, ym in enumerate(MONTHS)}  # 0..23

SELL_IN_SEASONALITY = {1: 0.85, 2: 1.15, 3: 1.25, 4: 1.15, 5: 0.90, 6: 0.80,
                        7: 0.75, 8: 1.05, 9: 1.30, 10: 1.20, 11: 0.85, 12: 0.75}
SELL_OUT_SEASONALITY = {1: 0.75, 2: 0.80, 3: 0.90, 4: 0.95, 5: 1.00, 6: 1.20,
                         7: 1.15, 8: 0.85, 9: 0.95, 10: 1.00, 11: 1.15, 12: 1.30}
DELIVERY_CANDIDATE_MONTHS = {2, 3, 4, 8, 9, 10}  # pre-book / seasonal shipment windows


def active_months_for_product(season):
    """Months (year, month) in which a product is part of the live assortment."""
    if season in ("SS24", "FW24"):
        return list(MONTHS)  # full 24-month window
    if season == "SS25":
        return [(y, m) for (y, m) in MONTHS if y == 2025]  # 12 months
    if season == "FW25":
        return [(y, m) for (y, m) in MONTHS if y == 2025 and m >= 7]  # 6 months
    raise ValueError(season)


# ---------------------------------------------------------------------------
# 4. ACCOUNT -> PRODUCT ASSORTMENT
# ---------------------------------------------------------------------------
def assortment_for_account(carried):
    footwear_slots = min(carried, len(FOOTWEAR_ORDER))
    leather_slots = carried - footwear_slots
    return FOOTWEAR_ORDER[:footwear_slots] + LEATHER_ORDER[:leather_slots]


for a in ACCOUNTS:
    a["products"] = assortment_for_account(a["carried"])

# ---------------------------------------------------------------------------
# 5. SELL-IN GENERATION
# ---------------------------------------------------------------------------
monthly_rows = []

for acct in ACCOUNTS:
    # revenue targets per year such that the two-year average == acct share * TARGET_ANNUAL_REVENUE
    base = acct["share"] * TARGET_ANNUAL_REVENUE
    trend = acct["trend"]
    rev_2024 = base * 2.0 / (1.0 + trend)
    rev_2025 = rev_2024 * trend
    year_revenue = {2024: rev_2024, 2025: rev_2025}

    n_products = len(acct["products"])
    delivery_k = 4 if acct["carried"] >= 13 else (3 if acct["carried"] >= 10 else random.choice([2, 3]))

    # per-account state: {product_id: {"stock": int}}
    stock_state = {pid: 0 for pid in acct["products"]}

    # pre-compute per-year product weights (active products only, weight = price * popularity)
    for year in (2024, 2025):
        active_products_year = [pid for pid in acct["products"]
                                 if (year, 1) in active_months_for_product(PRODUCTS_BY_ID[pid]["season"])
                                 or any(y == year for (y, m) in active_months_for_product(PRODUCTS_BY_ID[pid]["season"]))]
        weights = {}
        for pid in active_products_year:
            prod = PRODUCTS_BY_ID[pid]
            category_boost = 1.8 if prod["category"] == "Leather Goods" else 1.0
            weights[pid] = prod["wholesale"] * prod["popularity"] * category_boost
        wsum = sum(weights.values())
        for pid in active_products_year:
            weights[pid] /= wsum

        for pid in active_products_year:
            prod = PRODUCTS_BY_ID[pid]
            target_revenue = year_revenue[year] * weights[pid]
            target_units = target_revenue / prod["wholesale"]

            active_months_this_year = [(y, m) for (y, m) in active_months_for_product(prod["season"]) if y == year]
            candidates = sorted(set(m for (y, m) in active_months_this_year) & DELIVERY_CANDIDATE_MONTHS)
            if not candidates:
                candidates = sorted(set(m for (y, m) in active_months_this_year))[:1]
            k = min(delivery_k, len(candidates))
            delivery_months = set(random.sample(candidates, k))

            seas_weights = {m: SELL_IN_SEASONALITY[m] for m in delivery_months}
            sw_sum = sum(seas_weights.values())
            deliveries = {}
            running = 0
            dm_sorted = sorted(delivery_months)
            for i, m in enumerate(dm_sorted):
                noise = random.uniform(0.85, 1.15)
                units = target_units * (seas_weights[m] / sw_sum) * noise
                units = max(0, int(round(units)))
                deliveries[m] = units
                running += units

            # stash on product/year for the sell-through pass below
            acct.setdefault("_deliveries", {})[(pid, year)] = deliveries

    # ---- monthly walk: sell-in (from precomputed deliveries) + sell-through simulation ----
    for pid in acct["products"]:
        prod = PRODUCTS_BY_ID[pid]
        months_active = active_months_for_product(prod["season"])
        stock = 0
        for (year, month) in months_active:
            gidx = MONTH_INDEX[(year, month)]
            trend_frac = gidx / 23.0
            st_rate = acct["st_start"] + (acct["st_end"] - acct["st_start"]) * trend_frac
            st_rate *= SELL_OUT_SEASONALITY[month]
            st_rate = min(st_rate, 0.95)

            sell_in_units = acct["_deliveries"].get((pid, year), {}).get(month, 0)
            sell_in_value = sell_in_units * prod["wholesale"]

            available = stock + sell_in_units
            noise = random.uniform(0.85, 1.15)
            sell_out_units = int(round(available * st_rate * noise))
            sell_out_units = max(0, min(sell_out_units, available))
            sell_out_value = sell_out_units * prod["retail"]

            ret_noise = random.uniform(0.7, 1.3)
            returns_units = int(round(sell_out_units * acct["return_rate"] * ret_noise))
            returns_units = max(0, min(returns_units, sell_out_units))

            stock_end = max(0, available - sell_out_units - returns_units)

            monthly_rows.append(dict(
                account_id=acct["id"], product_id=pid,
                year=year, month=month,
                sell_in_units=sell_in_units, sell_in_value=sell_in_value,
                sell_out_units=sell_out_units, sell_out_value=sell_out_value,
                stock_units=stock_end, returns_units=returns_units,
            ))
            stock = stock_end

for acct in ACCOUNTS:
    acct.pop("_deliveries", None)

monthly_df = pd.DataFrame(monthly_rows)

# ---------------------------------------------------------------------------
# 6. VALIDATION / SANITY CHECK
# ---------------------------------------------------------------------------
total_sell_in = monthly_df["sell_in_value"].sum()
total_2024 = monthly_df.loc[monthly_df.year == 2024, "sell_in_value"].sum()
total_2025 = monthly_df.loc[monthly_df.year == 2025, "sell_in_value"].sum()
print(f"Total sell-in value (24 months): CHF {total_sell_in:,.0f}")
print(f"  2024: CHF {total_2024:,.0f}")
print(f"  2025: CHF {total_2025:,.0f}")
print(f"  Average per year: CHF {(total_sell_in / 2):,.0f}")

merged = monthly_df.merge(pd.DataFrame(PRODUCTS)[["id", "category"]], left_on="product_id", right_on="id")
cat_share = merged.groupby("category")["sell_in_value"].sum() / total_sell_in * 100
print("\nCategory share of sell-in value:")
print(cat_share.round(1))

# ---------------------------------------------------------------------------
# 7. WEEKLY SELL-OUT
# ---------------------------------------------------------------------------
def month_weeks(year, month):
    """ISO (iso_year, iso_week) tuples whose majority of days fall in (year, month)."""
    first = dt.date(year, month, 1)
    if month == 12:
        last = dt.date(year, 12, 31)
    else:
        last = dt.date(year, month + 1, 1) - dt.timedelta(days=1)

    day_counts = defaultdict(lambda: defaultdict(int))
    d = first
    while d <= last:
        iso_y, iso_w, _ = d.isocalendar()
        day_counts[(iso_y, iso_w)][(d.year, d.month)] += 1
        d += dt.timedelta(days=1)

    weeks = []
    for iso_key, month_counts in day_counts.items():
        majority_month = max(month_counts.items(), key=lambda kv: kv[1])[0]
        if majority_month == (year, month):
            weeks.append(iso_key)
    return sorted(weeks)


weekly_rows = []
for row in monthly_rows:
    if row["sell_out_units"] <= 0:
        continue
    weeks = month_weeks(row["year"], row["month"])
    n = len(weeks)
    if n == 0:
        continue
    raw_weights = [random.uniform(0.6, 1.4) for _ in range(n)]
    wsum = sum(raw_weights)
    weights = [w / wsum for w in raw_weights]

    unit_price = row["sell_out_value"] / row["sell_out_units"]
    exact_units = [row["sell_out_units"] * w for w in weights]
    floor_units = [int(u) for u in exact_units]
    remainder = row["sell_out_units"] - sum(floor_units)
    fracs = sorted(range(n), key=lambda i: exact_units[i] - floor_units[i], reverse=True)
    for i in fracs[:remainder]:
        floor_units[i] += 1

    for (iso_y, iso_w), units in zip(weeks, floor_units):
        if units <= 0:
            continue
        weekly_rows.append(dict(
            account_id=row["account_id"], product_id=row["product_id"],
            week=f"{iso_y}-W{iso_w:02d}",
            sell_out_units=units, sell_out_value=round(units * unit_price, 2),
        ))

weekly_df = pd.DataFrame(weekly_rows)
print(f"\nMonthly rows: {len(monthly_df):,}")
print(f"Weekly rows: {len(weekly_df):,}")
weekly_total = weekly_df["sell_out_value"].sum()
monthly_sellout_total = monthly_df["sell_out_value"].sum()
print(f"Sell-out value check -- monthly total: {monthly_sellout_total:,.0f} | weekly total: {weekly_total:,.0f}")

# ---------------------------------------------------------------------------
# 8. BUILD OUTPUT TABLES
# ---------------------------------------------------------------------------
accounts_df = pd.DataFrame([
    dict(**{"Account ID": a["id"], "Account Name": a["name"], "Region": a["region"],
            "Country": a["country"], "Account Type": a["acct_type"]})
    for a in ACCOUNTS
])

products_df = pd.DataFrame([
    dict(**{"Product ID": p["id"], "Product Name": p["name"], "Category": p["category"],
            "Season": p["season"], "Wholesale Price (CHF)": p["wholesale"], "Retail Price (CHF)": p["retail"]})
    for p in PRODUCTS
])

trans_monthly_df = pd.DataFrame([
    {
        "Account ID": r["account_id"], "Product ID": r["product_id"],
        "Month": f"{r['year']:04d}-{r['month']:02d}",
        "Sell-In Units": r["sell_in_units"], "Sell-In Value (CHF)": r["sell_in_value"],
        "Sell-Out Units": r["sell_out_units"], "Sell-Out Value (CHF)": r["sell_out_value"],
        "Stock Units": r["stock_units"], "Returns Units": r["returns_units"],
    }
    for r in monthly_rows
])

trans_weekly_df = pd.DataFrame([
    {
        "Account ID": r["account_id"], "Product ID": r["product_id"], "Week": r["week"],
        "Sell-Out Units": r["sell_out_units"], "Sell-Out Value (CHF)": r["sell_out_value"],
    }
    for r in weekly_rows
])
trans_weekly_df = trans_weekly_df.sort_values(["Week", "Account ID", "Product ID"]).reset_index(drop=True)
trans_monthly_df = trans_monthly_df.sort_values(["Month", "Account ID", "Product ID"]).reset_index(drop=True)

# ---------------------------------------------------------------------------
# 9. WRITE XLSX (clean, flat, no formulas / pivots / merges)
# ---------------------------------------------------------------------------
OUT_PATH = "Vasseur_Wholesale_Database_2024_2025.xlsx"

wb = Workbook()
wb.remove(wb.active)

sheets = {
    "Accounts": accounts_df,
    "Products": products_df,
    "Transactions_Monthly": trans_monthly_df,
    "Transactions_Weekly": trans_weekly_df,
}

for sheet_name, df in sheets.items():
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 34)

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
